from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def inspect(path: Path) -> None:
    workbook = load_workbook(path, data_only=True)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            values = [cell.value for cell in row]
            if any(isinstance(value, str) and ("询价失败" in value or "询价状态" in value) for value in values):
                trimmed = [value for value in values if value is not None]
                print(sheet.title, trimmed[:20])


def main() -> None:
    parser = argparse.ArgumentParser(description="Inspect market quote result rows.")
    parser.add_argument("path")
    args = parser.parse_args()
    inspect(Path(args.path))


if __name__ == "__main__":
    main()
