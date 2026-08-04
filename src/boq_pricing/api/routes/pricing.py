from __future__ import annotations

import shutil
import uuid
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import delete, select

from boq_pricing.api.dependencies import get_mysql_client, get_session_factory, get_settings
from boq_pricing.api.schemas import (
    PricingRunResponse,
    PricingRunDetail,
    PricingRunSummary,
    PricingResultSummary,
    PricingTaskAccepted,
    PricingTaskStatus,
)
from boq_pricing.application.async_tasks import PricingTaskRunner
from boq_pricing.application.batch import PricingBatchRequest, PricingBatchService
from boq_pricing.application.pricing_messages import build_no_active_rule_message
from boq_pricing.infrastructure import PricingTaskCreate, PricingTaskRepository, SqlAlchemyPriceRuleRepository
from boq_pricing.infrastructure.db import session_scope
from boq_pricing.infrastructure.orm_models import PricingResultORM, PricingRunORM, PricingTaskORM
from boq_pricing.pricing.calculations import calculate_total_price

router = APIRouter(prefix="/pricing", tags=["pricing"])


@router.post("/runs", response_model=PricingRunResponse)
def create_pricing_run(
    file: UploadFile = File(...),
    tenant_code: str = Form("default"),
    project_name: str | None = Form(None),
    region_code: str | None = Form(None),
    specialty: str | None = Form(None),
    cost_category: str | None = Form(None),
    rule_version: str | None = Form(None),
    write_mysql_audit: bool = Form(True),
) -> PricingRunResponse:
    settings = get_settings()
    session_factory = get_session_factory(settings)
    rules = SqlAlchemyPriceRuleRepository(session_factory, tenant_code=tenant_code).load(
        version=rule_version,
        region_code=region_code,
        specialty=specialty,
        cost_category=cost_category,
    )
    if not rules:
        raise HTTPException(
            status_code=422,
            detail=build_no_active_rule_message(rule_version, region_code, specialty, cost_category),
        )

    input_path = save_upload(file, settings.upload_dir)
    response = PricingBatchService().run(
        PricingBatchRequest(
            input_path=input_path,
            output_dir=settings.output_dir,
            rules=rules,
            rule_source="mysql",
            tenant_code=tenant_code,
            project_name=project_name,
            region_code=region_code,
            write_mysql_audit=write_mysql_audit,
            mysql_client=get_mysql_client(settings),
            session_factory=session_factory,
        )
    )
    return PricingRunResponse(
        item_count=response.item_count,
        priced_count=response.priced_count,
        unpriced_count=response.unpriced_count,
        issue_counts=response.issue_counts,
        excel_path=str(response.excel_path),
        missing_rules_path=str(response.missing_rules_path),
        audit_path=str(response.audit_path),
        mysql_run_code=response.mysql_run_code,
    )


@router.post("/tasks", response_model=PricingTaskAccepted, status_code=202)
def create_pricing_task(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    tenant_code: str = Form("default"),
    project_name: str | None = Form(None),
    region_code: str | None = Form(None),
    specialty: str | None = Form(None),
    cost_category: str | None = Form(None),
    rule_version: str | None = Form(None),
) -> PricingTaskAccepted:
    settings = get_settings()
    session_factory = get_session_factory(settings)
    task_code = uuid.uuid4().hex
    input_path = save_upload(file, settings.upload_dir, prefix=task_code)
    repository = PricingTaskRepository(session_factory)
    task = repository.create(
        PricingTaskCreate(
            tenant_code=tenant_code,
            task_code=task_code,
            workbook_name=file.filename or input_path.name,
            upload_path=str(input_path),
            project_name=project_name,
            region_code=region_code,
            specialty=specialty,
            cost_category=cost_category,
            rule_version=rule_version,
        )
    )
    runner = PricingTaskRunner(
        session_factory=session_factory,
        mysql_client=get_mysql_client(settings),
        output_dir=settings.output_dir,
    )
    background_tasks.add_task(runner.run, tenant_code, task_code)
    return PricingTaskAccepted(
        task_code=task.task_code,
        status=task.status,
        progress=task.progress,
        message=task.message,
    )


@router.get("/tasks", response_model=list[PricingTaskStatus])
def list_pricing_tasks(tenant_code: str = "default", limit: int = 20) -> list[PricingTaskStatus]:
    limit = max(1, min(limit, 100))
    tasks = PricingTaskRepository(get_session_factory()).list_recent(tenant_code, limit)
    return [to_task_status(task) for task in tasks]


@router.get("/tasks/{task_code}", response_model=PricingTaskStatus)
def get_pricing_task(task_code: str, tenant_code: str = "default") -> PricingTaskStatus:
    task = PricingTaskRepository(get_session_factory()).get(tenant_code, task_code)
    if task is None:
        raise HTTPException(status_code=404, detail="Pricing task was not found.")
    return to_task_status(task)


@router.post("/tasks/{task_code}/cancel", response_model=PricingTaskStatus)
def cancel_pricing_task(task_code: str, tenant_code: str = "default") -> PricingTaskStatus:
    repository = PricingTaskRepository(get_session_factory())
    task = repository.get(tenant_code, task_code)
    if task is None:
        raise HTTPException(status_code=404, detail="Pricing task was not found.")
    if task.status in {"succeeded", "failed", "canceled"}:
        return to_task_status(task)
    repository.mark_canceled(tenant_code, task_code)
    canceled = repository.get(tenant_code, task_code)
    if canceled is None:
        raise HTTPException(status_code=404, detail="Pricing task was not found.")
    return to_task_status(canceled)


@router.post("/tasks/{task_code}/resume", response_model=PricingTaskStatus)
def resume_pricing_task(
    task_code: str,
    background_tasks: BackgroundTasks,
    tenant_code: str = "default",
) -> PricingTaskStatus:
    settings = get_settings()
    session_factory = get_session_factory(settings)
    repository = PricingTaskRepository(session_factory)
    task = repository.get(tenant_code, task_code)
    if task is None:
        raise HTTPException(status_code=404, detail="Pricing task was not found.")
    if task.status not in {"canceled", "failed", "waiting_market_quote", "waiting_mapping"}:
        return to_task_status(task)
    repository.mark_pending(tenant_code, task_code)
    runner = PricingTaskRunner(
        session_factory=session_factory,
        mysql_client=get_mysql_client(settings),
        output_dir=settings.output_dir,
    )
    background_tasks.add_task(runner.run, tenant_code, task_code)
    resumed = repository.get(tenant_code, task_code)
    if resumed is None:
        raise HTTPException(status_code=404, detail="Pricing task was not found.")
    return to_task_status(resumed)


@router.get("/runs", response_model=list[PricingRunSummary])
def list_pricing_runs(tenant_code: str = "default", limit: int = 20) -> list[PricingRunSummary]:
    limit = max(1, min(limit, 100))
    with session_scope(get_session_factory()) as session:
        rows = session.scalars(
            select(PricingRunORM)
            .where(PricingRunORM.tenant_code == tenant_code)
            .order_by(PricingRunORM.id.desc())
            .limit(limit)
        ).all()
        return [
            PricingRunSummary(
                run_code=row.run_code,
                project_name=row.project_name,
                region_code=row.region_code,
                item_count=row.item_count,
                priced_count=row.priced_count,
                unpriced_count=row.unpriced_count,
                created_at=str(row.created_at),
                updated_at=str(row.updated_at),
            )
            for row in rows
        ]


@router.get("/runs/{run_code}", response_model=PricingRunDetail)
def get_pricing_run_detail(run_code: str, tenant_code: str = "default") -> PricingRunDetail:
    with session_scope(get_session_factory()) as session:
        run = session.scalar(
            select(PricingRunORM).where(
                PricingRunORM.tenant_code == tenant_code,
                PricingRunORM.run_code == run_code,
            )
        )
        if run is None:
            raise HTTPException(status_code=404, detail="Pricing run was not found.")
        results = session.scalars(
            select(PricingResultORM)
            .where(PricingResultORM.run_id == run.id)
            .order_by(PricingResultORM.source_sheet.asc(), PricingResultORM.source_row_number.asc())
        ).all()
        return PricingRunDetail(
            run_code=run.run_code,
            workbook_name=run.workbook_name,
            project_name=run.project_name,
            region_code=run.region_code,
            rule_source=run.rule_source,
            rule_version=run.rule_version,
            item_count=run.item_count,
            priced_count=run.priced_count,
            unpriced_count=run.unpriced_count,
            created_at=str(run.created_at),
            updated_at=str(run.updated_at),
            results=[to_result_summary(row) for row in results],
        )


@router.get("/runs/{run_code}/download")
def download_pricing_run_excel(run_code: str, tenant_code: str = "default") -> FileResponse:
    settings = get_settings()
    with session_scope(get_session_factory(settings)) as session:
        run = session.scalar(
            select(PricingRunORM).where(
                PricingRunORM.tenant_code == tenant_code,
                PricingRunORM.run_code == run_code,
            )
        )
        if run is None:
            raise HTTPException(status_code=404, detail="Pricing run was not found.")
        task = session.scalar(
            select(PricingTaskORM)
            .where(
                PricingTaskORM.tenant_code == tenant_code,
                PricingTaskORM.mysql_run_code == run_code,
            )
            .order_by(PricingTaskORM.id.desc())
        )
        candidate_paths = []
        if task and task.excel_path:
            candidate_paths.append(Path(task.excel_path))
        candidate_paths.append(settings.output_dir / f"{Path(run.workbook_name).stem}.priced.xlsx")

    excel_path = first_existing_excel(candidate_paths)
    if excel_path is None:
        raise HTTPException(status_code=404, detail="该批次没有可下载的结果 Excel，请确认任务已完成。")
    return FileResponse(
        path=excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{Path(excel_path).stem}.xlsx",
    )


@router.delete("/runs/{run_code}", status_code=204)
def delete_pricing_run(run_code: str, tenant_code: str = "default") -> None:
    with session_scope(get_session_factory()) as session:
        run = session.scalar(
            select(PricingRunORM).where(
                PricingRunORM.tenant_code == tenant_code,
                PricingRunORM.run_code == run_code,
            )
        )
        if run is None:
            raise HTTPException(status_code=404, detail="Pricing run was not found.")
        session.execute(
            delete(PricingTaskORM).where(
                PricingTaskORM.tenant_code == tenant_code,
                PricingTaskORM.mysql_run_code == run_code,
            )
        )
        session.delete(run)


def first_existing_excel(paths: list[Path]) -> Path | None:
    for path in paths:
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            continue
        if path.exists() and path.is_file():
            return path
    return None


def save_upload(file: UploadFile, upload_dir: Path, prefix: str | None = None) -> Path:
    upload_dir.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename or "boq.xlsx").name
    if prefix:
        safe_name = f"{prefix}.{safe_name}"
    destination = upload_dir / safe_name
    with destination.open("wb") as output:
        shutil.copyfileobj(file.file, output)
    return destination


def to_task_status(task) -> PricingTaskStatus:
    return PricingTaskStatus(
        task_code=task.task_code,
        status=task.status,
        progress=task.progress,
        message=task.message,
        workbook_name=task.workbook_name,
        project_name=task.project_name,
        region_code=task.region_code,
        item_count=task.item_count,
        priced_count=task.priced_count,
        unpriced_count=task.unpriced_count,
        excel_path=task.excel_path,
        missing_rules_path=task.missing_rules_path,
        audit_path=task.audit_path,
        mysql_run_code=task.mysql_run_code,
        failure_reasons=extract_task_failure_reasons(task.excel_path),
        created_at=str(task.created_at),
        started_at=str(task.started_at) if task.started_at else None,
        finished_at=str(task.finished_at) if task.finished_at else None,
    )


def to_result_summary(row: PricingResultORM) -> PricingResultSummary:
    total_price = calculate_total_price(row.quantity, row.unit_price)
    return PricingResultSummary(
        source_sheet=row.source_sheet,
        source_row_number=row.source_row_number,
        sequence_no=row.sequence_no,
        item_code=row.item_code,
        item_name=row.item_name,
        unit=row.unit,
        quantity=str(row.quantity) if row.quantity is not None else None,
        unit_price=str(row.unit_price) if row.unit_price is not None else None,
        total_price=str(total_price) if total_price is not None else None,
        rule_code=row.rule_code,
        rule_version=row.rule_version,
        price_source=row.price_source,
        confidence=str(row.confidence),
        features=dict(row.features_json or {}),
        issues=list(row.issues_json or []),
    )


def extract_task_failure_reasons(excel_path: str | None, limit: int = 10) -> list[str]:
    if not excel_path:
        return []
    path = Path(excel_path)
    if not path.exists() or path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return []
    reasons: list[str] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if not isinstance(value, str):
                        continue
                    if (
                        not value.startswith("询价失败：")
                        and "映射存在歧义" not in value
                        and "映射置信度不足" not in value
                    ):
                        continue
                    reason = value
                    if "映射存在歧义" in value:
                        reason = value
                    if reason not in reasons:
                        reasons.append(reason)
                    if len(reasons) >= limit:
                        return reasons
    except Exception:
        return []
    return reasons
