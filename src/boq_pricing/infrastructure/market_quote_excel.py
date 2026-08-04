from __future__ import annotations

import json
import re
import time
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from boq_pricing.infrastructure.excel import ExcelBillReader, find_header
from boq_pricing.infrastructure.market_quote_provider import (
    MarketQuoteProviderError,
    MarketQuoteRequest,
    MarketQuoteResult,
    unique_urls,
)
from boq_pricing.infrastructure.market_quotes import MarketQuoteRepository
from boq_pricing.infrastructure.model_call_logs import ModelCallLogRepository
from boq_pricing.infrastructure.source_verification import verify_source_url
from boq_pricing.parsing import FeatureParser

MONEY_QUANT = Decimal("0.01")


@dataclass(frozen=True)
class ExcelMarketQuoteSummary:
    item_count: int
    quoted_count: int
    failed_count: int
    output_path: str


class ExcelMarketQuoteService:
    def __init__(
        self,
        provider,
        repository: MarketQuoteRepository,
        output_dir: Path,
        call_log_repository: ModelCallLogRepository | None = None,
        task_code: str | None = None,
    ) -> None:
        self._provider = provider
        self._repository = repository
        self._output_dir = output_dir
        self._call_log_repository = call_log_repository
        self._task_code = task_code
        self._feature_parser = FeatureParser()

    def quote_workbook(
        self,
        input_path: Path,
        username: str,
        region: str | None = None,
        price_month: str | None = None,
        standard: str | None = None,
        limit: int = 50,
        progress_callback: Callable[[int, int, int, int, str], None] | None = None,
        cancel_checker: Callable[[], bool] | None = None,
    ) -> ExcelMarketQuoteSummary:
        items = ExcelBillReader().read(input_path)
        items = items[: max(1, min(limit, 200))]
        workbook = load_workbook(input_path)
        quoted_count = 0
        failed_count = 0

        quote_by_position: dict[tuple[str, int], list[str | float | None]] = {}
        for index, item in enumerate(items, start=1):
            if cancel_checker and cancel_checker():
                break
            features = self._feature_parser.parse(item.feature_text).values
            request = MarketQuoteRequest(
                item_name=item.item_name,
                unit=item.unit,
                features=features,
                region=region,
                price_month=price_month,
                standard=standard,
                item_code=item.item_code,
                quantity=str(item.quantity) if item.quantity is not None else None,
                work_content=item.work_content,
                remark=item.remark,
            )
            try:
                cached = self._repository.find_reusable(
                    item_name=item.item_name,
                    unit=item.unit,
                    features=features,
                    region_code=region,
                )
                if cached is not None:
                    quote_by_position[(item.source.sheet, item.source.row_number)] = build_cached_output_cells(
                        cached,
                        item.quantity,
                    )
                    quoted_count += 1
                    continue

                call_code = None
                started = time.monotonic()
                if self._call_log_repository is not None:
                    call_code = self._call_log_repository.start(
                        provider=self._provider.provider_name,
                        model=self._provider.model,
                        scenario="excel_market_quote",
                        task_code=self._task_code,
                        item_name=item.item_name,
                        username=username,
                    )
                try:
                    result = self._provider.quote_suppliers(request)
                    validate_supplier_quote_result(result)
                    if call_code is not None:
                        self._call_log_repository.succeed(
                            call_code,
                            duration_ms=int((time.monotonic() - started) * 1000),
                            token_usage=result.assumptions.get("token_usage"),
                            response_excerpt=result.raw_response,
                        )
                except MarketQuoteProviderError as exc:
                    if call_code is not None:
                        self._call_log_repository.fail(
                            call_code,
                            duration_ms=int((time.monotonic() - started) * 1000),
                            error_message=str(exc),
                        )
                    raise
                market_result = supplier_result_to_market_quote_result(result)
                average_unit_price = market_result.recommended_price
                self._repository.save(market_result, username=username)
                quote_by_position[(item.source.sheet, item.source.row_number)] = build_output_cells(
                    result,
                    item.quantity,
                    average_unit_price,
                    status="本次询价",
                )
                quoted_count += 1
            except MarketQuoteProviderError as exc:
                quote_by_position[(item.source.sheet, item.source.row_number)] = build_error_cells(str(exc))
                failed_count += 1
            finally:
                if progress_callback:
                    progress = 5 + int(index / max(len(items), 1) * 90)
                    progress_callback(
                        progress,
                        len(items),
                        quoted_count,
                        failed_count,
                        f"正在询价第 {index}/{len(items)} 行：{item.item_name}",
                    )

        append_quote_columns(workbook, quote_by_position)
        self._output_dir.mkdir(parents=True, exist_ok=True)
        output_path = self._output_dir / f"{input_path.stem}-市场询价结果.xlsx"
        workbook.save(output_path)
        return ExcelMarketQuoteSummary(
            item_count=len(items),
            quoted_count=quoted_count,
            failed_count=failed_count,
            output_path=str(output_path),
        )


def calculate_total_price(quantity: object, unit_price: Decimal | None) -> Decimal | None:
    if quantity is None or unit_price is None:
        return None
    return (Decimal(str(quantity)) * unit_price).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def normalize_money(value: Decimal | None) -> Decimal | None:
    if value is None:
        return None
    return value.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)


def average_price(quotes, require_source_url: bool = False, verified_source_urls: list[str] | None = None) -> Decimal | None:
    valid_quotes = [
        quote
        for quote in quotes
        if quote.price is not None and (not require_source_url or quote.source_url in set(verified_source_urls or []))
    ]
    if not valid_quotes:
        return None
    return (sum((quote.price for quote in valid_quotes), Decimal("0")) / Decimal(len(valid_quotes))).quantize(
        MONEY_QUANT,
        rounding=ROUND_HALF_UP,
    )


def min_price(quotes) -> Decimal | None:
    prices = [quote.price for quote in quotes if quote.price is not None]
    return min(prices) if prices else None


def max_price(quotes) -> Decimal | None:
    prices = [quote.price for quote in quotes if quote.price is not None]
    return max(prices) if prices else None

def validate_supplier_quote_result(result) -> None:
    require_source_url = bool(result.assumptions.get("require_source_url"))
    verified_source_urls: list[str] = []
    source_verifications: list[dict[str, object]] = []
    if require_source_url:
        source_verifications = verify_quote_source_urls(result)
        verified_source_urls = verified_urls_from_verifications(source_verifications)
        result.assumptions["source_verifications"] = source_verifications
        result.assumptions["verified_source_urls"] = verified_source_urls

    valid_quotes = valid_supplier_quotes(
        result,
        require_source_url=require_source_url,
        verified_source_urls=verified_source_urls,
        source_verifications=source_verifications,
    )
    result.assumptions["valid_quote_count"] = len(valid_quotes)
    result.assumptions["rejected_quote_count"] = max(0, len(result.quotes) - len(valid_quotes))
    result.assumptions["quote_quality_policy"] = (
        "有效报价必须同时具备明确价格、计价单位、可追溯来源链接，并且证据文本或来源页面需要出现该价格；"
        "仅出现产品名称但没有明确价格的页面不采纳。"
    )
    if len(valid_quotes) < 1:
        if require_source_url:
            raise MarketQuoteProviderError("询价失败：模型未返回带明确价格、可点击来源链接且通过页面证据校验的有效报价。")
        raise MarketQuoteProviderError("询价失败：模型未返回带明确价格和来源证据的有效报价。")
    if average_price(
        valid_quotes,
        require_source_url=require_source_url,
        verified_source_urls=verified_source_urls,
    ) is None:
        raise MarketQuoteProviderError("询价失败：模型返回的报价无法计算平均单价。")


def valid_supplier_quotes(
    result,
    require_source_url: bool = False,
    verified_source_urls: list[str] | None = None,
    source_verifications: list[dict[str, object]] | None = None,
):
    verified = set(verified_source_urls or [])
    verification_text_by_url = source_text_by_url(source_verifications or [])
    valid = []
    for quote in result.quotes:
        if quote.price is None or not quote.unit:
            continue
        if quote.evidence and re.search(r"(没有|无|未).*明确价格|没有.*价格", quote.evidence):
            continue
        if require_source_url and quote.source_url not in verified:
            continue
        if not require_source_url and not (quote.source_url or quote.evidence):
            continue
        evidence_text = " ".join(
            str(part or "")
            for part in (
                quote.evidence,
                quote.source_title,
                verification_text_by_url.get(quote.source_url or ""),
            )
        )
        if verification_text_by_url.get(quote.source_url or "") and not evidence_contains_price(evidence_text, quote.price):
            continue
        valid.append(quote)
    return valid


def source_text_by_url(source_verifications: list[dict[str, object]]) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in source_verifications:
        text = str(item.get("text_excerpt") or "")
        for key in ("url", "final_url"):
            url = str(item.get(key) or "")
            if url and text:
                result[url] = text
    return result


def evidence_contains_price(text: str, price: Decimal) -> bool:
    if not text:
        return False
    normalized_text = text.replace(",", "")
    candidates = {
        str(price.normalize()),
        format(price, "f").rstrip("0").rstrip("."),
        str(price.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)),
    }
    candidates = {item for item in candidates if item}
    return any(item in normalized_text for item in candidates)


def supplier_result_to_market_quote_result(result, average_unit_price: Decimal | None = None) -> MarketQuoteResult:
    require_source_url = bool(result.assumptions.get("require_source_url"))
    verified_source_urls = list(result.assumptions.get("verified_source_urls", []))
    valid_quotes = valid_supplier_quotes(
        result,
        require_source_url=require_source_url,
        verified_source_urls=verified_source_urls,
        source_verifications=list(result.assumptions.get("source_verifications", [])),
    )
    if average_unit_price is None:
        average_unit_price = average_price(
            valid_quotes,
            require_source_url=require_source_url,
            verified_source_urls=verified_source_urls,
        )
    return MarketQuoteResult(
        provider=result.provider,
        model=result.model,
        item_name=result.item_name,
        unit=result.unit,
        features=result.features,
        region_code=result.region_code,
        price_min=min_price(valid_quotes),
        price_max=max_price(valid_quotes),
        recommended_price=average_unit_price,
        tax_included=all(quote.tax_included for quote in valid_quotes) if valid_quotes else True,
        confidence=result.confidence,
        source_urls=verified_source_urls
        if require_source_url
        else unique_urls([*[quote.source_url for quote in valid_quotes if quote.source_url], *result.source_urls]),
        assumptions={
            **result.assumptions,
            "average_price_rule": "取所有通过证据校验的有效商家报价的算术平均值作为回填单价；只有一家有效报价时采用该报价。",
            "supplier_quotes": [
                {
                    "supplier": quote.supplier,
                    "brand": quote.brand,
                    "price": str(quote.price) if quote.price is not None else None,
                    "unit": quote.unit,
                    "source_url": quote.source_url,
                    "source_title": quote.source_title,
                    "evidence": quote.evidence,
                    "tax_included": quote.tax_included,
                }
                for quote in valid_quotes
            ],
        },
        raw_response=result.raw_response,
    )

def verify_quote_source_urls(result) -> list[dict[str, object]]:
    urls = unique_urls(
        [
            *[quote.source_url for quote in result.quotes if quote.source_url],
            *result.source_urls,
        ]
    )
    terms = build_quote_match_terms(result)
    verifications: list[dict[str, object]] = []
    for url in urls:
        verification = verify_source_url(url).to_dict()
        matched_terms = match_terms(str(verification.get("text_excerpt") or ""), terms)
        content_confirmed = True
        if verification.get("text_excerpt"):
            content_confirmed = bool(matched_terms)
        verification["content_confirmed"] = content_confirmed
        verification["matched_terms"] = matched_terms
        if not content_confirmed:
            verification["reason"] = verification.get("reason") or "页面内容未匹配清单名称、规格、品牌或价格关键词"
        verifications.append(verification)
    return verifications


def verified_urls_from_verifications(source_verifications: list[dict[str, object]]) -> list[str]:
    return unique_urls(
        [
            str(item.get("url") or "")
            for item in source_verifications
            if item.get("reachable") and item.get("specific_page", True) and item.get("content_confirmed", True)
        ]
        + [
            str(item.get("final_url") or "")
            for item in source_verifications
            if item.get("reachable") and item.get("specific_page", True) and item.get("content_confirmed", True)
        ]
    )


def build_quote_match_terms(result) -> list[str]:
    terms: list[str] = [result.item_name or "", result.unit or "", result.region_code or ""]
    terms.extend(str(value) for value in dict(result.features or {}).values() if value)
    for quote in result.quotes:
        terms.extend(
            [
                quote.supplier or "",
                quote.brand or "",
                quote.source_title or "",
                quote.evidence or "",
                str(quote.price) if quote.price is not None else "",
            ]
        )
    cleaned: list[str] = []
    for term in terms:
        value = str(term).strip()
        if len(value) >= 2 and value not in cleaned:
            cleaned.append(value)
    return cleaned[:40]


def match_terms(text: str, terms: list[str]) -> list[str]:
    if not text:
        return []
    lowered = text.lower()
    return [term for term in terms if term.lower() in lowered][:12]


def build_output_cells(result, quantity: object, average_unit_price: Decimal | None, status: str) -> list[str | float | None]:
    valid_quotes = valid_supplier_quotes(
        result,
        require_source_url=bool(result.assumptions.get("require_source_url")),
        verified_source_urls=list(result.assumptions.get("verified_source_urls", [])),
        source_verifications=list(result.assumptions.get("source_verifications", [])),
    )
    cells = [
        status,
        *supplier_cells(valid_quotes),
        str(result.confidence),
        "；".join(unique_urls([*[quote.source_url for quote in valid_quotes if quote.source_url], *result.source_urls])),
        json.dumps(result.assumptions, ensure_ascii=False),
    ]
    if average_unit_price is not None:
        cells[2] = float(average_unit_price)
    return cells


def build_cached_output_cells(cached, quantity: object) -> list[str | float | None]:
    supplier_quotes = []
    assumptions = getattr(cached, "assumptions_json", None) or {}
    for item in assumptions.get("supplier_quotes", [])[:3]:
        supplier_quotes.append(
            {
                "brand": item.get("brand"),
                "supplier": item.get("supplier"),
                "price": item.get("price"),
                "source_url": item.get("source_url"),
            }
        )
    return [
        "已询价复用",
        *supplier_dict_cells(supplier_quotes),
        str(getattr(cached, "confidence", "") or ""),
        "；".join(unique_urls([item.get("source_url") for item in supplier_quotes])),
        json.dumps(assumptions, ensure_ascii=False),
    ]


def build_error_cells(message: str) -> list[str | float | None]:
    return ["询价失败", None, None, None, None, None, None, None, None, None, message]


def supplier_cells(quotes) -> list[str | float | None]:
    cells: list[str | float | None] = []
    for quote in list(quotes)[:3]:
        cells.extend([
            quote.brand or quote.supplier,
            float(quote.price) if quote.price is not None else None,
            quote.source_url,
        ])
    while len(cells) < 9:
        cells.append(None)
    return cells


def supplier_dict_cells(quotes: list[dict[str, object]]) -> list[str | float | None]:
    cells: list[str | float | None] = []
    for quote in quotes[:3]:
        price = quote.get("price")
        cells.extend([
            str(quote.get("brand") or quote.get("supplier") or "") or None,
            float(price) if price not in (None, "") else None,
            str(quote.get("source_url") or "") or None,
        ])
    while len(cells) < 9:
        cells.append(None)
    return cells


def append_quote_columns(workbook, quote_by_position: dict[tuple[str, int], list[str | float | None]]) -> None:
    headers = [
        "询价状态",
        "询价品牌/供应商1",
        "询价单价1",
        "询价来源1",
        "询价品牌/供应商2",
        "询价单价2",
        "询价来源2",
        "询价品牌/供应商3",
        "询价单价3",
        "询价来源3",
        "询价说明",
    ]
    header_fill = PatternFill("solid", fgColor="E8F0EC")
    for sheet in workbook.worksheets:
        header_row, mapping = find_header(sheet)
        if header_row is None:
            continue
        start_col = sheet.max_column + 1
        for offset, header in enumerate(headers):
            cell = sheet.cell(row=header_row, column=start_col + offset, value=header)
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for (sheet_name, row_number), cells in quote_by_position.items():
            if sheet_name != sheet.title:
                continue
            status = cells[0] if cells else None
            unit_price = first_number(cells[2:9:3])
            quantity = sheet.cell(row=row_number, column=mapping["quantity"] + 1).value if "quantity" in mapping else None
            if unit_price is not None and "unit_price" in mapping:
                sheet.cell(row=row_number, column=mapping["unit_price"] + 1, value=unit_price)
            total_price = calculate_total_price(quantity, Decimal(str(unit_price))) if unit_price is not None else None
            if total_price is not None and "total_price" in mapping:
                sheet.cell(row=row_number, column=mapping["total_price"] + 1, value=float(total_price))
            for offset, value in enumerate(cells):
                cell = sheet.cell(row=row_number, column=start_col + offset, value=value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            if status == "询价失败":
                sheet.cell(row=row_number, column=start_col).fill = PatternFill("solid", fgColor="FEE4E2")


def first_number(values: list[str | float | None]) -> float | None:
    for value in values:
        if value is None or value == "":
            continue
        try:
            return float(value)
        except (TypeError, ValueError):
            continue
    return None


