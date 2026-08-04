from __future__ import annotations

from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from boq_pricing.domain import BillItem, PricingResult, SourceRef
from boq_pricing.pricing.calculations import calculate_total_price


HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "sequence": ("序号",),
    "item_code": ("项目编码", "清单编码"),
    "item_name": ("项目名称", "清单名称"),
    "feature_text": ("项目特征", "项目特征描述"),
    "unit": ("计量单位", "单位"),
    "quantity": ("工程量", "数量"),
    "unit_price": ("综合单价", "单价"),
    "total_price": ("合价", "金额"),
    "work_content": ("工作内容",),
    "remark": ("备注",),
}

HEADER_ALIASES = {
    "sequence": ("序号", "编号"),
    "item_code": ("项目编码", "清单编码"),
    "item_name": ("项目名称", "清单名称"),
    "feature_text": ("项目特征", "项目特征描述", "技术指标", "技术指标（规格型号）", "技术指标(规格型号)"),
    "unit": ("计量单位", "单位"),
    "quantity": ("工程量", "数量"),
    "unit_price": ("综合单价", "单价", "单价（元）", "单价(元)", "价格"),
    "total_price": ("合价", "金额", "合价（元）", "合价(元)"),
    "work_content": ("工作内容",),
    "remark": ("备注",),
}


class ExcelBillReader:
    def read(self, workbook_path: str | Path) -> list[BillItem]:
        path = Path(workbook_path)
        workbook = load_workbook(path, data_only=True, read_only=False)
        items: list[BillItem] = []

        for sheet in workbook.worksheets:
            header_row, mapping = find_header(sheet)
            if header_row is None:
                continue
            for row_number, row_cells in enumerate(
                sheet.iter_rows(min_row=header_row + 1, values_only=False),
                start=header_row + 1,
            ):
                row = tuple(cell.value for cell in row_cells)
                item = build_item(path.name, sheet.title, row_number, row, mapping)
                if item:
                    apply_display_precision(item, row_cells, mapping)
                    items.append(item)
        return items


def find_header(sheet: Worksheet) -> tuple[int | None, dict[str, int]]:
    for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        mapping = map_headers(row)
        if "item_name" in mapping and ("quantity" in mapping or "total_price" in mapping):
            return row_number, mapping
    return None, {}


def map_headers(row: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized = [normalize_cell(value) for value in row]
    for field, aliases in HEADER_ALIASES.items():
        for index, cell in enumerate(normalized):
            if cell in aliases:
                mapping[field] = index
                break
    return mapping


def build_item(
    workbook_name: str,
    sheet_name: str,
    row_number: int,
    row: tuple[Any, ...],
    mapping: dict[str, int],
) -> BillItem | None:
    item_name = get_cell(row, mapping, "item_name")
    quantity = to_decimal(get_cell(row, mapping, "quantity"))
    unit = normalize_cell(get_cell(row, mapping, "unit")) or None

    if not item_name:
        return None
    if quantity is None and not unit:
        return None

    return BillItem(
        sequence=to_text(get_cell(row, mapping, "sequence")),
        item_code=to_text(get_cell(row, mapping, "item_code")),
        item_name=normalize_cell(item_name),
        feature_text=to_text(get_cell(row, mapping, "feature_text")) or "",
        unit=unit,
        quantity=quantity,
        original_unit_price=to_decimal(get_cell(row, mapping, "unit_price")),
        original_total_price=to_decimal(get_cell(row, mapping, "total_price")),
        work_content=to_text(get_cell(row, mapping, "work_content")),
        remark=to_text(get_cell(row, mapping, "remark")),
        source=SourceRef(workbook=workbook_name, sheet=sheet_name, row_number=row_number),
    )


def get_cell(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> Any:
    index = mapping.get(field)
    if index is None or index >= len(row):
        return None
    return row[index]


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\n", " ").strip()


def to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def apply_display_precision(item: BillItem, row_cells: tuple[Any, ...], mapping: dict[str, int]) -> None:
    unit_price_cell = get_cell(row_cells, mapping, "unit_price")
    total_price_cell = get_cell(row_cells, mapping, "total_price")
    item.original_unit_price = to_decimal_from_displayed_number(unit_price_cell, item.original_unit_price)
    item.original_total_price = to_decimal_from_displayed_number(total_price_cell, item.original_total_price)


def to_decimal_from_displayed_number(cell: Any, fallback: Decimal | None) -> Decimal | None:
    if fallback is None or cell is None:
        return fallback
    places = decimal_places_from_number_format(getattr(cell, "number_format", ""))
    if places is None:
        return fallback
    quant = Decimal("1").scaleb(-places)
    return fallback.quantize(quant, rounding=ROUND_HALF_UP)


def decimal_places_from_number_format(number_format: str | None) -> int | None:
    if not number_format:
        return None
    fmt = str(number_format).split(";")[0]
    if fmt.lower() == "general" or "." not in fmt:
        return None
    decimal_part = fmt.split(".", 1)[1]
    places = 0
    index = 0
    while index < len(decimal_part):
        char = decimal_part[index]
        if char in "0#?":
            places += 1
            index += 1
            continue
        if char == "\\":
            index += 2
            continue
        if char in "_*":
            index += 2
            continue
        break
    return places if places > 0 else None


class ExcelResultWriter:
    def write(
        self,
        results: list[PricingResult],
        output_path: str | Path,
        source_path: str | Path | None = None,
    ) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        if source_path is not None:
            return self.write_into_source_workbook(results, path, Path(source_path))

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "计价结果"

        headers = [
            "Sheet",
            "行号",
            "序号",
            "项目编码",
            "项目名称",
            "计量单位",
            "工程量",
            "综合单价",
            "合价",
            "单价来源",
            "规则ID",
            "置信度",
            "项目特征JSON",
            "异常提示",
        ]
        sheet.append(headers)
        for result in results:
            item = result.item
            quote = result.quote
            total_price = calculated_total_price(result)
            sheet.append(
                [
                    item.source.sheet,
                    item.source.row_number,
                    item.sequence,
                    item.item_code,
                    item.item_name,
                    item.unit,
                    float(item.quantity) if item.quantity is not None else None,
                    float(quote.unit_price) if quote.unit_price is not None else None,
                    float(total_price) if total_price is not None else None,
                    quote.source,
                    quote.rule_id,
                    quote.confidence,
                    str(item.features.values if item.features else {}),
                    "; ".join(issue.message for issue in result.issues),
                ]
            )

        style_result_sheet(sheet)
        workbook.save(path)
        return path

    def write_into_source_workbook(
        self,
        results: list[PricingResult],
        output_path: Path,
        source_path: Path,
    ) -> Path:
        workbook = load_workbook(source_path)
        result_by_position = {
            (result.item.source.sheet, result.item.source.row_number): result
            for result in results
        }
        for sheet in workbook.worksheets:
            header_row, mapping = find_header(sheet)
            if header_row is None:
                continue
            row_numbers = sorted(
                row_number
                for sheet_name, row_number in result_by_position
                if sheet_name == sheet.title
            )
            if not row_numbers:
                continue
            start_column = logical_last_column(sheet, header_row) + 1
            append_pricing_headers(sheet, header_row, start_column)
            for row_number in row_numbers:
                result = result_by_position[(sheet.title, row_number)]
                write_pricing_result_to_original_row(sheet, row_number, mapping, result)
                append_pricing_result_cells(sheet, row_number, start_column, result)

        workbook.save(output_path)
        return output_path


class ExcelMissingRuleTemplateWriter:
    def write(self, results: list[PricingResult], output_path: str | Path) -> Path:
        path = Path(output_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "待补价格规则"
        headers = [
            "rule_code",
            "version",
            "item_name_contains",
            "unit",
            "feature_conditions_json",
            "unit_price",
            "source",
            "sample_count",
            "sample_sheet",
            "sample_row",
            "sample_feature_text",
        ]
        sheet.append(headers)

        grouped: dict[tuple[str, str | None, str], dict[str, Any]] = {}
        for result in results:
            if result.quote.unit_price is not None:
                continue
            item = result.item
            features = item.features.values if item.features else {}
            feature_json = json.dumps(features, ensure_ascii=False, sort_keys=True)
            key = (item.item_name, item.unit, feature_json)
            if key not in grouped:
                grouped[key] = {
                    "item": item,
                    "features": features,
                    "count": 0,
                }
            grouped[key]["count"] += 1

        for index, record in enumerate(grouped.values(), start=1):
            item = record["item"]
            sheet.append(
                [
                    f"AUTO-{index:04d}",
                    "2026.07-draft",
                    item.item_name,
                    item.unit,
                    json.dumps(record["features"], ensure_ascii=False),
                    None,
                    "待填写",
                    record["count"],
                    item.source.sheet,
                    item.source.row_number,
                    item.feature_text,
                ]
            )

        style_missing_rule_sheet(sheet)
        workbook.save(path)
        return path


class ExcelPriceRuleReader:
    def read(self, workbook_path: str | Path) -> list[dict[str, Any]]:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [normalize_cell(value) for value in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {
                headers[index]: row[index]
                for index in range(min(len(headers), len(row)))
                if headers[index]
            }
            if record.get("rule_code") and record.get("unit_price") not in (None, ""):
                records.append(record)
        return records


def logical_last_column(sheet: Worksheet, header_row: int) -> int:
    for column in range(sheet.max_column, 0, -1):
        if sheet.cell(header_row, column).value not in (None, ""):
            return column
    return sheet.max_column


def append_pricing_headers(sheet: Worksheet, header_row: int, start_column: int) -> None:
    headers = [
        "清单识别",
        "特征摘要",
        "取价状态",
        "计价综合单价",
        "计价合价",
        "计价规则版本",
        "计价规则ID",
        "取价来源",
        "置信度",
        "匹配条件",
        "计价说明",
    ]
    template = sheet.cell(header_row, max(1, start_column - 1))
    for offset, header in enumerate(headers):
        cell = sheet.cell(row=header_row, column=start_column + offset, value=header)
        copy_cell_style(template, cell)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = 18 if offset not in {1, 9, 10} else 36


def write_pricing_result_to_original_row(
    sheet: Worksheet,
    row_number: int,
    mapping: dict[str, int],
    result: PricingResult,
) -> None:
    unit_price_column = mapping.get("unit_price")
    total_price_column = mapping.get("total_price")
    total_price = calculated_total_price(result)
    if unit_price_column is not None and result.quote.unit_price is not None:
        cell = sheet.cell(row=row_number, column=unit_price_column + 1, value=float(result.quote.unit_price))
        cell.number_format = "0.00"
    if total_price_column is not None and total_price is not None:
        cell = sheet.cell(row=row_number, column=total_price_column + 1, value=float(total_price))
        cell.number_format = "0.00"


def append_pricing_result_cells(
    sheet: Worksheet,
    row_number: int,
    start_column: int,
    result: PricingResult,
) -> None:
    quote = result.quote
    total_price = calculated_total_price(result)
    values: list[str | float | None] = [
        format_item_identity(result),
        feature_summary(result),
        "已取价" if quote.unit_price is not None else "待补价",
        float(quote.unit_price) if quote.unit_price is not None else None,
        float(total_price) if total_price is not None else None,
        quote.rule_version,
        quote.rule_id,
        quote.source,
        quote.confidence,
        json.dumps(quote.matched_conditions, ensure_ascii=False),
        "; ".join(issue.message for issue in result.issues),
    ]
    template = sheet.cell(row_number, max(1, start_column - 1))
    for offset, value in enumerate(values):
        cell = sheet.cell(row=row_number, column=start_column + offset, value=value)
        copy_cell_style(template, cell)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if offset in {3, 4} and value is not None:
            cell.number_format = "0.00"


def calculated_total_price(result: PricingResult) -> Decimal | None:
    return calculate_total_price(result.item.quantity, result.quote.unit_price)


def format_item_identity(result: PricingResult) -> str:
    item = result.item
    parts = [f"Sheet: {item.source.sheet}", f"行号: {item.source.row_number}"]
    if item.sequence:
        parts.append(f"序号: {item.sequence}")
    if item.item_code:
        parts.append(f"编码: {item.item_code}")
    return "；".join(parts)


def feature_summary(result: PricingResult, limit: int = 140) -> str:
    features = result.item.features.values if result.item.features else {}
    if features:
        text = "；".join(f"{key}={value}" for key, value in features.items())
    else:
        text = result.item.feature_text or ""
    return text[:limit] + ("..." if len(text) > limit else "")


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def style_result_sheet(sheet: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "A": 34,
        "B": 8,
        "C": 10,
        "D": 16,
        "E": 28,
        "F": 10,
        "G": 14,
        "H": 14,
        "I": 16,
        "J": 24,
        "K": 24,
        "L": 10,
        "M": 48,
        "N": 48,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def style_missing_rule_sheet(sheet: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="806000")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    widths = {
        "A": 16,
        "B": 16,
        "C": 28,
        "D": 10,
        "E": 56,
        "F": 14,
        "G": 24,
        "H": 12,
        "I": 34,
        "J": 10,
        "K": 64,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
