from decimal import Decimal
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from boq_pricing.infrastructure import market_quote_excel
from boq_pricing.infrastructure.market_quote_excel import ExcelMarketQuoteService
from boq_pricing.infrastructure.market_quote_excel import calculate_total_price
from boq_pricing.infrastructure.market_quote_excel import validate_supplier_quote_result
from boq_pricing.infrastructure.market_quote_provider import MarketQuoteProviderError
from boq_pricing.infrastructure.market_quote_provider import SupplierMarketQuote, SupplierMarketQuoteResult
from boq_pricing.infrastructure.source_verification import SourceVerification


class FakeSupplierQuoteProvider:
    def __init__(self):
        self.requests = []

    def quote_suppliers(self, request):
        self.requests.append(request)
        return SupplierMarketQuoteResult(
            provider="fake",
            model="fake-model",
            item_name=request.item_name,
            unit=request.unit,
            features=request.features,
            region_code=request.region,
            recommended_price=Decimal("2.50"),
            confidence=Decimal("0.88"),
            quotes=[
                SupplierMarketQuote("供应商A", "品牌A", Decimal("2.40"), "m", "https://a.example", "报价A", "规格匹配", True),
                SupplierMarketQuote("供应商B", "品牌B", Decimal("2.50"), "m", "https://b.example", "报价B", "规格匹配", True),
                SupplierMarketQuote("供应商C", "品牌C", Decimal("2.60"), "m", "https://c.example", "报价C", "规格匹配", True),
            ],
            source_urls=["https://a.example", "https://b.example", "https://c.example"],
            assumptions={"口径": "含税不含安装"},
            raw_response='{"ok": true}',
        )


class FakeMarketQuoteRepository:
    def __init__(self):
        self.saved = []
        self.reusable = None

    def save(self, result, username=None):
        self.saved.append((result, username))

    def find_reusable(self, item_name, unit, features, region_code):
        return self.reusable


def test_excel_market_quote_service_fills_unit_price_total_and_sources(tmp_path):
    workbook_path = tmp_path / "询价表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单"
    sheet.append(["分部分项工程量清单与计价表"])
    sheet.append(["序号", "项目编码", "项目名称", "项目特征", "计量单位", "工程量", "综合单价", "合价", "工作内容", "备注"])
    sheet.append(["1.1", None, "预制钢筋混凝土桩", "1.桩型：PHC-300-AB-70\n2.桩长：8-10m\n3.混凝土：C80", "m", 4, None, None, "桩体运输、压桩", ""])
    workbook.save(workbook_path)

    provider = FakeSupplierQuoteProvider()
    repository = FakeMarketQuoteRepository()
    summary = ExcelMarketQuoteService(provider, repository, tmp_path).quote_workbook(
        workbook_path,
        username="tester",
        region="CN",
        price_month="2026-07",
        standard="GB13476-2023",
    )

    output = load_workbook(summary.output_path, data_only=True)
    output_sheet = output["清单"]

    assert summary.item_count == 1
    assert summary.quoted_count == 1
    assert summary.failed_count == 0
    assert output_sheet["G3"].value == 2.5
    assert output_sheet["H3"].value == 10
    assert output_sheet["K2"].value == "询价状态"
    assert output_sheet["K3"].value == "本次询价"
    assert output_sheet["L2"].value == "询价品牌/供应商1"
    assert output_sheet["N3"].value == "https://a.example"
    assert provider.requests[0].work_content == "桩体运输、压桩"
    assert provider.requests[0].features["桩长度"] == "8-10m"
    assert len(repository.saved) == 1


def test_excel_market_quote_service_accepts_one_valid_supplier_quote(tmp_path):
    workbook_path = tmp_path / "单家询价表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单"
    sheet.append(["序号", "项目编码", "项目名称", "项目特征", "计量单位", "工程量", "综合单价", "合价", "工作内容", "备注"])
    sheet.append(["1.1", None, "预制钢筋混凝土桩", "桩型：PHC-300-AB-70", "m", 4, None, None, "压桩", ""])
    workbook.save(workbook_path)

    class OneQuoteProvider:
        def quote_suppliers(self, request):
            return SupplierMarketQuoteResult(
                provider="fake",
                model="fake-model",
                item_name=request.item_name,
                unit=request.unit,
                features=request.features,
                region_code=request.region,
                recommended_price=Decimal("2.50"),
                confidence=Decimal("0.70"),
                quotes=[
                    SupplierMarketQuote("供应商A", "品牌A", Decimal("2.50"), "m", "https://a.example", "报价A", "规格匹配", True)
                ],
                source_urls=["https://a.example"],
                assumptions={"有效报价数量": "1"},
                raw_response='{"ok": true}',
            )

    repository = FakeMarketQuoteRepository()
    summary = ExcelMarketQuoteService(OneQuoteProvider(), repository, tmp_path).quote_workbook(
        workbook_path,
        username="tester",
    )

    output = load_workbook(summary.output_path, data_only=True)
    output_sheet = output["清单"]

    assert summary.quoted_count == 1
    assert summary.failed_count == 0
    assert output_sheet["G2"].value == 2.5
    assert output_sheet["H2"].value == 10
    assert output_sheet["K1"].value == "询价状态"
    assert output_sheet["L1"].value == "询价品牌/供应商1"
    assert output_sheet["O1"].value == "询价品牌/供应商2"


def test_excel_market_quote_service_reuses_existing_quote_and_skips_provider(tmp_path):
    workbook_path = tmp_path / "复用询价表.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "清单"
    sheet.append(["序号", "项目编码", "项目名称", "项目特征", "计量单位", "工程量", "综合单价", "合价", "工作内容", "备注"])
    sheet.append(["1.1", None, "预制钢筋混凝土桩", "桩型：PHC-300-AB-70", "m", 4, None, None, "压桩", ""])
    workbook.save(workbook_path)

    class ProviderShouldNotRun:
        def __init__(self):
            self.calls = 0

        def quote_suppliers(self, request):
            self.calls += 1
            raise AssertionError("cached quote should skip model provider")

    provider = ProviderShouldNotRun()
    repository = FakeMarketQuoteRepository()
    repository.reusable = SimpleNamespace(
        recommended_price=Decimal("2.50"),
        confidence=Decimal("0.80"),
        quote_code="cached-001",
        created_at="2026-07-22 10:00:00",
        assumptions_json={
            "supplier_quotes": [
                {
                    "supplier": "历史供应商",
                    "brand": "历史品牌",
                    "price": "2.50",
                    "unit": "m",
                    "source_url": "https://cached.example",
                    "evidence": "历史询价来源",
                }
            ]
        },
        raw_response='{"cached": true}',
    )

    summary = ExcelMarketQuoteService(provider, repository, tmp_path).quote_workbook(
        workbook_path,
        username="tester",
        region="CN",
    )

    output = load_workbook(summary.output_path, data_only=True)
    output_sheet = output["清单"]

    assert summary.quoted_count == 1
    assert summary.failed_count == 0
    assert provider.calls == 0
    assert repository.saved == []
    assert output_sheet["G2"].value == 2.5
    assert output_sheet["H2"].value == 10
    assert output_sheet["K2"].value == "已询价复用"
    assert output_sheet["L2"].value == "历史品牌"
    assert output_sheet["N2"].value == "https://cached.example"


def test_calculate_total_price_uses_quantity_times_unit_price():
    assert calculate_total_price(Decimal("247180"), Decimal("88.00")) == Decimal("21751840.00")
    assert calculate_total_price(Decimal("129"), Decimal("1234.567")) == Decimal("159259.14")


def test_web_search_quote_requires_clickable_source_url():
    result = SupplierMarketQuoteResult(
        provider="fake",
        model="fake-model",
        item_name="水尺",
        unit="m",
        features={"材质": "不锈钢"},
        region_code="CN",
        recommended_price=Decimal("2.50"),
        confidence=Decimal("0.70"),
        quotes=[
            SupplierMarketQuote("供应商A", "品牌A", Decimal("2.50"), "m", None, "报价A", "只有文字证据", True)
        ],
        source_urls=[],
        assumptions={"require_source_url": True},
        raw_response='{"ok": true}',
    )

    try:
        validate_supplier_quote_result(result)
    except MarketQuoteProviderError as exc:
        assert "可点击来源链接" in str(exc)
    else:
        raise AssertionError("web search quotes without source_url must be rejected")


def test_web_search_quote_rejects_unreachable_source_url(monkeypatch):
    monkeypatch.setattr(
        market_quote_excel,
        "verify_source_url",
        lambda url: SourceVerification(url=url, reachable=False, reason="timeout"),
    )
    result = SupplierMarketQuoteResult(
        provider="fake",
        model="fake-model",
        item_name="水尺",
        unit="m",
        features={"材质": "不锈钢"},
        region_code="CN",
        recommended_price=Decimal("2.50"),
        confidence=Decimal("0.70"),
        quotes=[
            SupplierMarketQuote(
                "供应商A",
                "品牌A",
                Decimal("2.50"),
                "m",
                "https://bad.example/item",
                "报价A",
                "规格匹配",
                True,
            )
        ],
        source_urls=["https://bad.example/item"],
        assumptions={"require_source_url": True},
        raw_response='{"ok": true}',
    )

    try:
        validate_supplier_quote_result(result)
    except MarketQuoteProviderError as exc:
        assert "可点击来源链接" in str(exc)
        assert result.assumptions["source_verifications"][0]["reachable"] is False
    else:
        raise AssertionError("unreachable source_url must be rejected")


def test_web_search_quote_accepts_reachable_source_url(monkeypatch):
    monkeypatch.setattr(
        market_quote_excel,
        "verify_source_url",
        lambda url: SourceVerification(
            url=url,
            reachable=True,
            status_code=200,
            final_url="https://supplier.example/final",
            content_type="text/html",
        ),
    )
    result = SupplierMarketQuoteResult(
        provider="fake",
        model="fake-model",
        item_name="水尺",
        unit="m",
        features={"材质": "不锈钢"},
        region_code="CN",
        recommended_price=Decimal("2.50"),
        confidence=Decimal("0.70"),
        quotes=[
            SupplierMarketQuote(
                "供应商A",
                "品牌A",
                Decimal("2.50"),
                "m",
                "https://supplier.example/item",
                "报价A",
                "规格匹配",
                True,
            )
        ],
        source_urls=["https://supplier.example/item"],
        assumptions={"require_source_url": True},
        raw_response='{"ok": true}',
    )

    validate_supplier_quote_result(result)

    assert "https://supplier.example/item" in result.assumptions["verified_source_urls"]
    assert "https://supplier.example/final" in result.assumptions["verified_source_urls"]
