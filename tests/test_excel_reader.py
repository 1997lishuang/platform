from pathlib import Path
import unittest
from decimal import Decimal

from openpyxl import Workbook, load_workbook
from boq_pricing.domain import BillItem, PriceQuote, PricingResult, SourceRef
from boq_pricing.infrastructure import ExcelBillReader, ExcelResultWriter


class ExcelReaderTest(unittest.TestCase):
    def test_reads_real_workbook(self):
        workbook = Path("工程招标工程量清单.xlsx")
        items = ExcelBillReader().read(workbook)

        self.assertGreaterEqual(len(items), 100)
        self.assertTrue(any(item.item_name == "预制钢筋混凝土桩" for item in items))
        self.assertTrue(any(item.source.sheet.startswith("2-2") for item in items))

    def test_reads_unit_price_using_excel_display_precision(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["编号", "项目名称", "技术指标（规格型号）", "单位", "工程量", "单价（元）", "合价（元）"])
        sheet.append(["1", "镀锌钢管", "规格:DN80×t3.25", "m", 90, 58.347563625, 5251.28072625])
        sheet["F2"].number_format = "0.00_ "
        sheet["G2"].number_format = "0.00_ "
        path = Path("uploads/test-display-price.xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(path)

        items = ExcelBillReader().read(path)

        self.assertEqual(items[0].original_unit_price, Decimal("58.35"))
        self.assertEqual(items[0].original_total_price, Decimal("5251.28"))


if __name__ == "__main__":
    unittest.main()


def test_result_writer_appends_pricing_columns_to_each_source_sheet(tmp_path):
    workbook_path = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "安装工程"
    second = workbook.create_sheet("土建工程")
    headers = ["编号", "项目名称", "技术指标（规格型号）", "单位", "工程量", "单价（元）", "合价（元）"]
    first.append(headers)
    first.append(["1", "水尺", "材质:不锈钢", "m", 2, None, None])
    second.append(headers)
    second.append(["2", "镀锌钢管", "规格:DN80", "m", 3, None, None])
    workbook.save(workbook_path)

    results = [
        PricingResult(
            item=BillItem(
                sequence="1",
                item_code=None,
                item_name="水尺",
                feature_text="材质:不锈钢",
                unit="m",
                quantity=Decimal("2"),
                original_unit_price=None,
                original_total_price=None,
                work_content=None,
                remark=None,
                source=SourceRef(workbook_path.name, "安装工程", 2),
            ),
            quote=PriceQuote(
                unit_price=Decimal("25"),
                total_price=Decimal("999"),
                rule_id="R-1",
                rule_version="v1",
                source="unit-test",
                confidence=1.0,
                matched_conditions={},
            ),
        ),
        PricingResult(
            item=BillItem(
                sequence="2",
                item_code=None,
                item_name="镀锌钢管",
                feature_text="规格:DN80",
                unit="m",
                quantity=Decimal("3"),
                original_unit_price=None,
                original_total_price=None,
                work_content=None,
                remark=None,
                source=SourceRef(workbook_path.name, "土建工程", 2),
            ),
            quote=PriceQuote(
                unit_price=Decimal("58.35"),
                total_price=Decimal("175.05"),
                rule_id="R-2",
                rule_version="v1",
                source="unit-test",
                confidence=0.9,
                matched_conditions={"规格": "DN80"},
            ),
        ),
    ]

    output_path = tmp_path / "multi-sheet.priced.xlsx"
    ExcelResultWriter().write(results, output_path, source_path=workbook_path)

    output = load_workbook(output_path, data_only=True)
    assert output.sheetnames == ["安装工程", "土建工程"]
    assert output["安装工程"]["F2"].value == 25
    assert output["安装工程"]["G2"].value == 50
    assert output["安装工程"]["H1"].value == "清单识别"
    assert output["安装工程"]["I1"].value == "特征摘要"
    assert output["安装工程"]["J1"].value == "取价状态"
    assert output["安装工程"]["J2"].value == "已取价"
    assert output["安装工程"]["L1"].value == "计价合价"
    assert output["安装工程"]["L2"].value == 50
    assert output["土建工程"]["F2"].value == 58.35
    assert output["土建工程"]["G2"].value == 175.05
    assert output["土建工程"]["H1"].value == "清单识别"
